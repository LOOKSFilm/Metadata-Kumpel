from EditShareAPI import FlowMetadata
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.styles.differential import DifferentialStyle
from tkinter import filedialog
from tkinter import messagebox
import win32com.client as win32
import xml.etree.ElementTree as et
import customtkinter

def build(app, selects, assignfields, is_imported, xmlfile):
    #rows = int(dpg.get_value(item="rows"))
    path = filedialog.asksaveasfilename(title="Save Excel to...",defaultextension=".xlsm", filetypes=[("Excel", ".xlsm")], initialfile="metadata_template")
    workbook = Workbook()
    worksheet1 = workbook.active
    worksheet1.title = "Assign Metadata"
    worksheet2 = workbook.create_sheet(title="Data")
    column = 1
    worksheet1.freeze_panes = "A2"
    c1 = worksheet1.cell(row=1, column=column, value="Mapping ID")
    c1.font = Font(bold=True)
    c1.fill = PatternFill(fill_type="solid", start_color="00FF9900")
    c1.border = Border(bottom=Side(border_style="medium", color="00808080"))
    dvcount = 0
    multiselectfiels = list()
    oldvalues = list()
    
    size, x_coord, y_coord = app.geometry().split("+")
    
    font = customtkinter.CTkFont(family="Hack NF", size=12, weight="bold")
    window_loading = customtkinter.CTkToplevel()
    window_loading.title("Building Excel")
    window_loading.geometry(f"{400}x{70}+{str(int(x_coord)+100)}+{str(int(y_coord)+300)}")
    window_loading.deiconify()
    label_loading = customtkinter.CTkLabel(window_loading, text="Building Excel", font=font)
    label_loading.pack()
    progress_bar = customtkinter.CTkProgressBar(window_loading, mode="indeterminte", width=350)
    progress_bar.pack()
    progress_bar.start()

    for select in selects:
        column += 1
        c2 = worksheet1.cell(row=1, column=column, value=select)
        c2.font = Font(bold=True)
        c2.fill = PatternFill(fill_type="solid", start_color="00FFCC99")
        c2.border = Border(bottom=Side(border_style="medium", color="00808080"))
        field_id = assignfields[select]
        fielddata = FlowMetadata.getCustomMetadataField(field_id)
        fieldtype = fielddata["type"]
        multiselect = fielddata["multi_select"]

        if is_imported and not xmlfile == "":
            tree = et.parse(xmlfile)
            root = tree.findall("clip")
            root += tree.findall("image")
            for row, clip in enumerate(root):
                row = row + 2
                #print(clip)
                try:
                    clipname = clip.find("clipname").text
                except:
                    clipname = clip.find("file/backup/path").text.split("/")[-1].split(".")[0]
                    print(clipname)
                if not "003b Mapping Identifier" in selects:
                    worksheet1.cell(row=row, column=1).value = clipname
                metadata = clip.find("custom")
                for entry in metadata:
                    try:
                        field_from_xml = entry.attrib["username"]
                    except KeyError:
                        continue
                    if field_from_xml == select:
                        if select == "003b Mapping Identifier" and entry.text != "":
                            worksheet1.cell(row=row, column=1).value = entry.text
                        else:
                            worksheet1.cell(row=row, column=1).value = clipname
                        #print(field_from_xml, select)
                        worksheet1.cell(row=row, column=column).value = entry.text
                        #print(entry.text)
                        if fieldtype == "bool":
                            if worksheet1.cell(row=row, column=column).value == "false":
                                worksheet1.cell(row=row, column=column).value = "False"
                            else:
                                worksheet1.cell(row=row, column=column).value = "True"
                            dv = DataValidation(type="list", formula1='"True,False"', allow_blank=False)
                            rule1 = CellIsRule(operator="equal", formula=['"False"'], stopIfTrue=True, fill=PatternFill(start_color="00E07777", end_color="00E07777", fill_type="solid"), font=Font(bold=True))
                            rule2 = CellIsRule(operator="equal", formula=['"True"'], stopIfTrue=True, fill=PatternFill(start_color="0068D070",end_color="0068D070", fill_type="solid"), font=Font(bold=True))
                            worksheet1.conditional_formatting.add(worksheet1.cell(row=row, column=column).coordinate, rule1)
                            worksheet1.conditional_formatting.add(worksheet1.cell(row=row, column=column).coordinate, rule2)
                            worksheet1.add_data_validation(dv)
                            dv.add(worksheet1.cell(row=row, column=column))
                            #worksheet1.cell(row=row, column=column).value = "True"
                            #worksheet1.cell(row=row, column=column).fill = PatternFill(fill_type="solid", start_color="00FF733D")
                            #worksheet1.cell(row=row, column=column).fill = PatternFill(fill_type="solid", start_color="00CCFFFF")
                            worksheet1.cell(row=row, column=column).border = Border(bottom=Side(border_style="medium", color="00808080"), top=Side(border_style="medium", color="00808080"), left=Side(border_style="medium", color="00808080"), right=Side(border_style="medium", color="00808080"))
                        if fieldtype == "int":
                            dv = DataValidation(type="whole")
                            worksheet1.add_data_validation(dv)
                            dv.add(worksheet1.cell(row=row, column=column))
                            worksheet1.cell(row=row, column=column).fill = PatternFill(fill_type="solid", start_color="00CCFFFF")
                            worksheet1.cell(row=row, column=column).border = Border(bottom=Side(border_style="medium", color="00808080"), top=Side(border_style="medium", color="00808080"), left=Side(border_style="medium", color="00808080"), right=Side(border_style="medium", color="00808080"))
                        if fieldtype == "Qdate":
                            dv = DataValidation(type="date")
                            worksheet1.add_data_validation(dv)
                            dv.add(worksheet1.cell(row=row, column=column))
                            worksheet1.cell(row=row, column=column).fill = PatternFill(fill_type="solid", start_color="00CCFFFF")
                            worksheet1.cell(row=row, column=column).border = Border(bottom=Side(border_style="medium", color="00808080"), top=Side(border_style="medium", color="00808080"), left=Side(border_style="medium", color="00808080"), right=Side(border_style="medium", color="00808080"))
                        try:
                            values = fielddata["allowed_values"]["values"]
                            if not oldvalues == values:
                                dvcount += 1
                                for i, value in enumerate(values):
                                    d1 = worksheet2.cell(row=i+1, column=dvcount)
                                    d1.value = value["value"]
                                startcell = "$"+worksheet2.cell(row=1, column=dvcount).column_letter+"$"+str(worksheet2.cell(row=1, column=dvcount).row)
                                endcell = "$"+worksheet2.cell(row=len(values), column=dvcount).column_letter+"$"+str(worksheet2.cell(row=len(values), column=dvcount).row)
                                if multiselect:
                                    multiselectfiels.append(worksheet1.cell(row=row, column=column).column)
                                #print(startcell, endcell)
                                oldvalues = values
                            dv = DataValidation(type="list", formula1="Data!"+startcell+":"+endcell, allow_blank=True)
                            worksheet1.add_data_validation(dv)
                            dv.add(worksheet1.cell(row=row, column=column))
                            worksheet1.cell(row=row, column=column).fill = PatternFill(fill_type="solid", start_color="00CCFFFF")
                            worksheet1.cell(row=row, column=column).border = Border(bottom=Side(border_style="medium", color="00808080"), top=Side(border_style="medium", color="00808080"), left=Side(border_style="medium", color="00808080"), right=Side(border_style="medium", color="00808080"))
                                    
                        except KeyError:
                            pass                             
        else:
            row = 2
            if fieldtype == "bool":
                rule1 = CellIsRule(operator="equal", formula=['"False"'], stopIfTrue=True, fill=PatternFill(start_color="00E07777", end_color="00E07777", fill_type="solid"), font=Font(bold=True))
                rule2 = CellIsRule(operator="equal", formula=['"True"'], stopIfTrue=True, fill=PatternFill(start_color="0068D070",end_color="0068D070", fill_type="solid"), font=Font(bold=True))
                worksheet1.conditional_formatting.add(worksheet1.cell(row=row, column=column).coordinate, rule1)
                worksheet1.conditional_formatting.add(worksheet1.cell(row=row, column=column).coordinate, rule2)
                dv = DataValidation(type="list", formula1='"True,False"', allow_blank=False)
                worksheet1.add_data_validation(dv)
                dv.add(worksheet1.cell(row=row, column=column))
                #worksheet1.cell(row=row, column=column).value = "True"
                worksheet1.cell(row=row, column=column).fill = PatternFill(fill_type="solid", start_color="00CCFFFF")
                worksheet1.cell(row=row, column=column).border = Border(bottom=Side(border_style="medium", color="00808080"), top=Side(border_style="medium", color="00808080"), left=Side(border_style="medium", color="00808080"), right=Side(border_style="medium", color="00808080"))
            if fieldtype == "int":
                dv = DataValidation(type="whole")
                worksheet1.add_data_validation(dv)
                dv.add(worksheet1.cell(row=row, column=column))
                worksheet1.cell(row=row, column=column).fill = PatternFill(fill_type="solid", start_color="00CCFFFF")
                worksheet1.cell(row=row, column=column).border = Border(bottom=Side(border_style="medium", color="00808080"), top=Side(border_style="medium", color="00808080"), left=Side(border_style="medium", color="00808080"), right=Side(border_style="medium", color="00808080"))
            if fieldtype == "Qdate":
                dv = DataValidation(type="date")
                worksheet1.add_data_validation(dv)
                dv.add(worksheet1.cell(row=row, column=column))
                worksheet1.cell(row=row, column=column).fill = PatternFill(fill_type="solid", start_color="00CCFFFF")
                worksheet1.cell(row=row, column=column).border = Border(bottom=Side(border_style="medium", color="00808080"), top=Side(border_style="medium", color="00808080"), left=Side(border_style="medium", color="00808080"), right=Side(border_style="medium", color="00808080"))
            try:
                values = fielddata["allowed_values"]["values"]
                dvcount += 1
                for i, value in enumerate(values):
                    d1 = worksheet2.cell(row=i+1, column=dvcount)
                    d1.value = value["value"]
                startcell = "$"+worksheet2.cell(row=1, column=dvcount).column_letter+"$"+str(worksheet2.cell(row=1, column=dvcount).row)
                endcell = "$"+worksheet2.cell(row=len(values), column=dvcount).column_letter+"$"+str(worksheet2.cell(row=len(values), column=dvcount).row)
                #print(startcell, endcell)
                dv = DataValidation(type="list", formula1="Data!"+startcell+":"+endcell, allow_blank=True)
                worksheet1.add_data_validation(dv)
                dv.add(worksheet1.cell(row=row, column=column))
                worksheet1.cell(row=row, column=column).fill = PatternFill(fill_type="solid", start_color="00CCFFFF")
                worksheet1.cell(row=row, column=column).border = Border(bottom=Side(border_style="medium", color="00808080"), top=Side(border_style="medium", color="00808080"), left=Side(border_style="medium", color="00808080"), right=Side(border_style="medium", color="00808080"))
            except KeyError:
                pass
        if multiselect:
            multiselectfiels.append(worksheet1.cell(row=row, column=column).column)

    for c, multiselectfield in enumerate(multiselectfiels):
        if c == 0:
            codefields = f'If Target.Column = {multiselectfield}'
        else:
            codefields += f' Or Target.Column = {multiselectfield}'
        if c == len(multiselectfiels)-1:
            codefields += " Then"

    worksheet2.protection.sheet = True
    try:
        workbook.save(path)
        progress_bar.stop()
        window_loading.destroy()
        wb = load_workbook(filename=path, read_only=False, keep_vba=True)
        wb.save(path)
        xl = win32.gencache.EnsureDispatch('Excel.Application')
        xl.Visible = True
        ss = xl.Workbooks.Open(path)
        if multiselectfiels:
            xlmodule = ss.VBProject.VBComponents("Tabelle1")
            xlmodule.Name = 'multiselect'
            code = f'''Private Sub Worksheet_Change(ByVal Target As Range)
    Dim Oldvalue As String
    Dim Newvalue As String
    Application.EnableEvents = True
    On Error GoTo Exitsub
    {codefields}
        If Target.SpecialCells(xlCellTypeAllValidation) Is Nothing Then
            GoTo Exitsub
        Else
            If Target.Value = "" Then
                GoTo Exitsub
            Else
                Application.EnableEvents = False
                Newvalue = Target.Value
                Application.Undo
                Oldvalue = Target.Value
                If Oldvalue = "" Then
                    Target.Value = Newvalue
                Else
                    If InStr(1, Oldvalue, Newvalue) = 0 Then
                        Target.Value = Oldvalue & "; " & Newvalue
                    Else
                        ' Check if the Newvalue already exists in the list (deactivate)
                        Dim valuesArray() As String
                        valuesArray = Split(Oldvalue, "; ")
                        Dim modifiedValue As String
                        modifiedValue = ""
                        Dim isFirst As Boolean
                        isFirst = True
                        Dim i As Long
                        For i = LBound(valuesArray) To UBound(valuesArray)
                            If valuesArray(i) <> Newvalue Then
                                If Not isFirst Then
                                    modifiedValue = modifiedValue & "; "
                                End If
                                modifiedValue = modifiedValue & valuesArray(i)
                                isFirst = False
                            End If
                        Next i
                        Target.Value = modifiedValue
                    End If
                End If
            End If
        End If
    End If
    Application.EnableEvents = True
Exitsub:
    Application.EnableEvents = True
End Sub
'''
            xlmodule.CodeModule.AddFromString(code)

    except PermissionError:
        messagebox.showerror("Permission Error", "Can not save Excel file:\n\nClose the Excel Document!\n\n")
        pass
    #     if dpg.does_item_exist("Error"):
    #         dpg.delete_item("Error")
    #     with dpg.window(popup=True, tag="Error", pos=(400, 300)):
    #         dpg.bind_item_theme("Error", "errortheme")
    #         dpg.add_text(default_value="Can not save Excel file:\n\nClose the Excel Document!\n\n")
    #         dpg.add_button(label="OK", callback=lambda: dpg.delete_item(item="Error"), width=100)
